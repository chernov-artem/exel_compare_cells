import re
from openpyxl import load_workbook

""" Работа с exel файлами
Сравнивает 2 ячейки exel. Находитв первой ячейки модель товара и ищет в другой ячейке такую же модель. Пробелы и точки игнорируются.

"""

def only_digits(str: str) -> str:
    "оставляем в строке только цифры"
    res = re.sub("[^0-9]","", str)
    print(res)
    return res

def find_model_num(s1:str, s2: str) -> bool:
    "находит в строке1 последовательность из 2-15 цифрт и ищет его в строке2"
    list1 = []
    for i in s1.split(' '):
        if re.findall('\d{2,15}', i) != []:
            list1.append(i)
    print(list1)
    return only_digits(list1[0]) in only_digits(s2)

def compare_cells(num: int) -> bool:
    "получает на вход номер ряда и сравнивает ячейки B и C в этом ряду. Возвращает True в случае совпадения"
    num_cell_B = 'B' + str(num)
    num_cell_C = 'C' + str(num)
    tmp1 = ws[num_cell_B].value
    tmp2 = ws[num_cell_C].value
    print(tmp1)
    print(tmp2)
    if tmp1 != [] and tmp2 != []:
        try:
            print(tmp1.split(' ')[0], tmp2.split(' ')[0])
            print(find_model_num(tmp1, tmp2))
            return True
        except Exception as ex:
            print("ex = ", ex)
            return False

wb = load_workbook('sber18122023.xlsx')
ws = wb['Аня']


if __name__ == '__main__':
    print('my result = ', compare_cells(1200))